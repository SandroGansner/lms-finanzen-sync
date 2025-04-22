
import os
import requests
import pandas as pd
import schedule
import time
from datetime import datetime
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Font
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
import pickle
import img2pdf

# Lade die Umgebungsvariablen aus der .env-Datei
load_dotenv()

# Hole die Werte aus der .env-Datei
SUPABASE_URL = os.getenv("SUPABASE_URL")
API_KEY = os.getenv("API_KEY")

# Debugging: Überprüfe, ob die Umgebungsvariablen geladen wurden
print(f"SUPABASE_URL: {SUPABASE_URL}")
print(f"API_KEY: {API_KEY}")

# Setze die Header für die Anfrage
headers = {
    "apikey": API_KEY,
    "Authorization": f"Bearer {API_KEY}",
    "Content-Type": "application/json"
}

# Google Drive API-Einstellungen
SCOPES = ['https://www.googleapis.com/auth/drive']
CREDENTIALS_FILE = 'credentials.json'
TOKEN_FILE = 'token_purchases.pickle'

# Funktion zum Einrichten des Google Drive-Dienstes
def get_drive_service():
    creds = None
    if os.path.exists(TOKEN_FILE):
        with open(TOKEN_FILE, 'rb') as token:
            creds = pickle.load(token)
    if not creds or not creds.valid:
        flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_FILE, SCOPES)
        creds = flow.run_local_server(port=0)
        with open(TOKEN_FILE, 'wb') as token:
            pickle.dump(creds, token)
    return build('drive', 'v3', credentials=creds)

# Funktion zum Erstellen oder Finden eines Ordners in Google Drive
def get_or_create_folder(drive_service, folder_name, parent_id=None):
    query = f"name='{folder_name}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
    if parent_id:
        query += f" and '{parent_id}' in parents"
    response = drive_service.files().list(q=query, spaces='drive').execute()
    folders = response.get('files', [])
    if folders:
        return folders[0]['id']
    else:
        file_metadata = {
            'name': folder_name,
            'mimeType': 'application/vnd.google-apps.folder'
        }
        if parent_id:
            file_metadata['parents'] = [parent_id]
        folder = drive_service.files().create(body=file_metadata, fields='id').execute()
        return folder.get('id')

# Funktion zum Prüfen, ob eine Datei in Google Drive existiert
def file_exists_in_drive(drive_service, file_name, folder_id):
    query = f"name='{file_name}' and '{folder_id}' in parents and trashed=false"
    response = drive_service.files().list(q=query, spaces='drive').execute()
    files = response.get('files', [])
    return len(files) > 0

# Funktion zum Hochladen einer Datei nach Google Drive
def upload_to_drive(drive_service, file_path, file_name, folder_id):
    if file_exists_in_drive(drive_service, file_name, folder_id):
        print(f"Datei {file_name} existiert bereits in Google Drive, überspringe Upload.")
        return
    try:
        file_metadata = {
            'name': file_name,
            'parents': [folder_id]
        }
        media = MediaFileUpload(file_path)
        file = drive_service.files().create(body=file_metadata, media_body=media, fields='id').execute()
        print(f"Datei erfolgreich hochgeladen nach Google Drive: {file_name} (ID: {file.get('id')})")
    except Exception as e:
        print(f"Fehler beim Hochladen der Datei {file_name} nach Google Drive: {e}")

# Funktion zum Abrufen von Daten aus einer Supabase-Tabelle
def fetch_data(table_name):
    try:
        response = requests.get(f"{SUPABASE_URL}/rest/v1/{table_name}", headers=headers)
        if response.status_code == 200:
            return response.json()
        else:
            print(f"Fehler beim Abrufen von {table_name}: {response.status_code} - {response.text}")
            return []
    except Exception as e:
        print(f"Fehler beim Abrufen von {table_name}: {e}")
        return []

# Funktion zum Herunterladen eines Belegs und ggf. Umwandeln in PDF
def download_receipt(receipt_path, local_image_path, local_pdf_path):
    try:
        if not receipt_path:
            print("Kein receiptPath angegeben.")
            return False
        receipt_url = f"{SUPABASE_URL}/storage/v1/object/{receipt_path}"
        print(f"Versuche, Beleg herunterzuladen von: {receipt_url}")
        response = requests.get(receipt_url, headers=headers)
        if response.status_code == 200:
            # Prüfe, ob die Datei bereits ein PDF ist
            if receipt_path.lower().endswith('.pdf'):
                # Speichere das PDF direkt
                with open(local_pdf_path, 'wb') as f:
                    f.write(response.content)
                print(f"PDF-Beleg heruntergeladen: {local_pdf_path}")
                return True
            else:
                # Speichere das Bild temporär
                with open(local_image_path, 'wb') as f:
                    f.write(response.content)
                print(f"Bild heruntergeladen: {local_image_path}")
                # Wandle das Bild in ein PDF um
                with open(local_pdf_path, 'wb') as f:
                    f.write(img2pdf.convert(local_image_path))
                print(f"Bild in PDF umgewandelt: {local_pdf_path}")
                # Lösche das temporäre Bild
                os.remove(local_image_path)
                return True
        else:
            print(f"Fehler beim Herunterladen des Belegs {receipt_path}: {response.status_code} - {response.text}")
            return False
    except Exception as e:
        print(f"Fehler beim Herunterladen oder Umwandeln des Belegs {receipt_path}: {e}")
        return False

# Funktion zur Synchronisation der Einkäufe
def sync_purchases():
    print(f"Starte Synchronisation der Einkäufe: {datetime.now()}")
    purchases = fetch_data("purchases")
    if not purchases:
        print("Keine Einkäufe gefunden.")
        return

    # Konvertiere die Daten in ein DataFrame
    df_purchases = pd.DataFrame(purchases)
    if df_purchases.empty:
        print("DataFrame ist leer.")
        return

    # Konvertiere das created_date_time-Feld und erstelle eine Spalte für Monat/Jahr
    df_purchases['created_date_time'] = pd.to_datetime(df_purchases['created_date_time'])
    df_purchases['month_year'] = df_purchases['created_date_time'].dt.strftime('%Y_%m')

    # Initialisiere Google Drive-Dienst
    drive_service = get_drive_service()

    # Erstelle die Hauptordner in Google Drive
    einkaufe_folder_id = get_or_create_folder(drive_service, "Einkäufe")
    belege_folder_id = get_or_create_folder(drive_service, "Belege")
    belege_einkaufe_folder_id = get_or_create_folder(drive_service, "Einkäufe", belege_folder_id)

    # Gruppiere nach Kreditkarte und Monat (basierend auf created_date_time)
    grouped = df_purchases.groupby(['cardUsed', 'month_year'])

    # Erstelle oder aktualisiere Excel-Dateien und Belege-Ordner
    for (card, month_year), group in grouped:
        # Erstelle den Ordner für die Excel-Dateien (lokal und in Google Drive)
        excel_dir = f"exports/Einkäufe/{card.replace(' ', '_')}/{month_year}"
        os.makedirs(excel_dir, exist_ok=True)
        card_folder_id = get_or_create_folder(drive_service, card.replace(' ', '_'), einkaufe_folder_id)
        month_folder_id = get_or_create_folder(drive_service, month_year, card_folder_id)

        # Erstelle den Ordner für die Belege (lokal und in Google Drive)
        receipts_dir = f"exports/Belege/Einkäufe/{month_year}/Belege_{card.replace(' ', '_')}_{month_year}"
        os.makedirs(receipts_dir, exist_ok=True)
        belege_month_folder_id = get_or_create_folder(drive_service, month_year, belege_einkaufe_folder_id)
        belege_card_folder_id = get_or_create_folder(drive_service, f"Belege_{card.replace(' ', '_')}_{month_year}", belege_month_folder_id)

        # Wähle die relevanten Spalten für die Excel-Datei
        excel_data = group[[
            'id', 'invoiceIssuer', 'itemName', 'account', 'kst', 'project', 'vatRate', 'price'
        ]].copy()
        excel_data.insert(1, 'Beleg', '')
        excel_data['BETRAG EUR'] = ''
        excel_data.columns = [
            'ID', 'Beleg', 'Rechnungssteller', 'Text', 'Kontierung Konto', 'KST', 'Projekt', 'VAT', 'BETRAG CHF', 'BETRAG EUR'
        ]

        # Überprüfe, ob die Excel-Datei bereits existiert
        filename = f"{excel_dir}/Einkauf_{card.replace(' ', '_')}_{month_year}.xlsx"
        if os.path.exists(filename):
            try:
                existing_data = pd.read_excel(filename, skiprows=9)
                if not existing_data.empty and existing_data.iloc[-1]['ID'] == 'TOTAL':
                    existing_data = existing_data.iloc[:-1]
                excel_data = pd.concat([existing_data, excel_data], ignore_index=True)
                excel_data = excel_data.drop_duplicates(subset=['ID'], keep='last')
            except Exception as e:
                print(f"Fehler beim Lesen der bestehenden Excel-Datei {filename}: {e}")

        # Berechne die Summen
        sum_row = excel_data[['BETRAG CHF']].sum()
        sum_row['ID'] = 'TOTAL'
        sum_row['Beleg'] = ''
        sum_row['Rechnungssteller'] = ''
        sum_row['Text'] = ''
        sum_row['Kontierung Konto'] = ''
        sum_row['KST'] = ''
        sum_row['Projekt'] = ''
        sum_row['VAT'] = ''
        sum_row['BETRAG EUR'] = ''

        # Füge die Summenzeile hinzu
        excel_data = pd.concat([excel_data, pd.DataFrame([sum_row])], ignore_index=True)

        # Speichere die Excel-Datei
        try:
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                header_data = pd.DataFrame([
                    ['Kreditkarten-Abrechnung'],
                    [f"Anbieter: {card}"],
                    [f"Periode: {month_year.replace('_', ' ')}"],
                    ['Lenzerheide Marketing+Support AG'],
                    ['GJ 2024/25']
                ])
                header_data.to_excel(writer, sheet_name='Sheet1', startrow=0, index=False, header=False)
                excel_data.to_excel(writer, sheet_name='Sheet1', startrow=9, index=False)

            workbook = load_workbook(filename)
            worksheet = workbook['Sheet1']
            column_widths = {
                'A': 10, 'B': 10, 'C': 20, 'D': 30, 'E': 15, 'F': 10, 'G': 10, 'H': 10, 'I': 15, 'J': 15
            }
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width

            last_row = worksheet.max_row
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=last_row, column=col)
                cell.font = Font(bold=True)

            workbook.save(filename)
            print(f"Excel-Datei erstellt/aktualisiert: {filename}")
            upload_to_drive(drive_service, filename, os.path.basename(filename), month_folder_id)

        except Exception as e:
            print(f"Fehler beim Erstellen/Aktualisieren der Excel-Datei {filename}: {e}")

        # Lade die Belege herunter, wandle sie in PDF um und lade sie nach Google Drive hoch
        for index, row in group.iterrows():
            receipt_path = row['receiptPath']
            if receipt_path:
                receipt_id = row['id']
                item_name = row['itemName'].replace(' ', '_')[:20]
                # Temporärer Pfad für das Bild
                temp_image_path = f"{receipts_dir}/temp_{receipt_id}_{item_name}.jpg"
                # Endgültiger Pfad für das PDF
                new_filename = f"Beleg_{receipt_id}_{item_name}.pdf"
                local_pdf_path = f"{receipts_dir}/{new_filename}"
                if download_receipt(receipt_path, temp_image_path, local_pdf_path):
                    upload_to_drive(drive_service, local_pdf_path, new_filename, belege_card_folder_id)

# Hauptfunktion zur Synchronisation
def sync_all():
    sync_purchases()

# Plane die Synchronisation täglich um 2:00 Uhr
schedule.every().day.at("02:00").do(sync_all)

# Teste die Synchronisation sofort beim Start
sync_all()

# Starte den Scheduler
print("Starte Synchronisation... Drücke Ctrl+C zum Beenden.")
while True:
    schedule.run_pending()
    time.sleep(1)
EOF